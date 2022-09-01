import pandas as pd
import openpyxl as xl
import shutil
import datetime

OUTPUT_FILEPATH = "output_files/output_data/output_data.xlsx"
INVOICE_TEMPLATES  = "static/invoice_templates/invoice_template1.xlsx"
OUTPUT_INVOICES_DIRPATH = "output_files/output_invoices"

def main():
    # 請求書作成を実行
    make_invoice()

# エクセルを読み込む関数
def read_excel(filepath):
    df = pd.read_excel(filepath, index_col=0)
    return df

# データ上にある全ての顧客名（会社名）のリストを取得する（重複を消す）
def make_company_list():
    df = pd.read_excel(OUTPUT_FILEPATH)
    company_list = []
    company_list = list(df["顧客名"].unique())
    return company_list

# 各顧客別（会社別）の注文データを取得
def get_company_data(company):
    df=read_excel(OUTPUT_FILEPATH)
    df=df[df["顧客名"]==company]
    return df

# 請求書雛形を顧客の人数分にコピーして、名前を請求書名に変更
def copy_invoice_templates(company):
    shutil.copyfile(INVOICE_TEMPLATES, OUTPUT_INVOICES_DIRPATH + "/" + "【請求書】" + company + " 御中" +".xlsx")

# 顧客名と請求Noが入った辞書を作成関数{"あひる商事":"20220901-1"}
def make_number_dic(company_list):
    company_count = len(company_list)
    today = str(datetime.date.today().strftime('%Y%m%d'))

    # リスト生成[0,1,2,3,4・・・・]
    list = [x for x in range(1, company_count+1)] 

    # 辞書生成
    number_dic = {}
    for youso in list:
        number = today + "-" + str(list[youso-1])
        number_dic[company_list[youso-1]] = number
    return number_dic

# 請求書作成関数
def make_invoice():
    company_list = make_company_list()
    number_dic = make_number_dic(company_list)

    for company in company_list:
        df = get_company_data(company)
        copy_invoice_templates(company)
        wb = xl.load_workbook(OUTPUT_INVOICES_DIRPATH + "/" + "【請求書】" + company + " 御中" + ".xlsx")
        ws = wb.worksheets[0]

        # 会社名入れ込み
        ws["A2"].value = company

        # 請求日入れ込み
        ws["G3"].value = str(datetime.date.today())

        # ナンバー入れ込み （今日の日付-1）
        ws["G2"].value = number_dic[company]

        # 商品名入れ込み
        each_df = df[df["顧客名"]==company]
        each_df = each_df[["商品名", "数量", "単価"]].groupby("商品名").sum()
        product_list = each_df.index.values

        if len(product_list) > 14:
            print("14個以上の商品が登録されています。請求書を作成できません")
            pass
        else:
            i = 15
            for product in product_list:
                ws[f"A{i}"].value = product
                i = i+1

        # 数量入れ込み
        suryo_list = each_df["数量"].to_list()
        if len(suryo_list) > 14:
            print("14個以上の商品が登録されています。請求書を作成できません")
            pass
        else:
            i = 15
            for suryo in suryo_list:
                ws[f"D{i}"].value = suryo
                i = i+1

        # 単価入れ込み
        tanka_list = each_df["単価"].to_list()
        if len(tanka_list) > 14:
            print("14個以上の商品が登録されています。請求書を作成できません")
            pass
        else:     
            i = 15
            for tanka in tanka_list:
                ws[f"F{i}"].value = tanka
                i = i+1
        # ブック保存
        wb.save(OUTPUT_INVOICES_DIRPATH + "/" + "【請求書】" + company + " 御中" + ".xlsx")
        
# メイン関数の実行
if __name__ == "__main__":
    main()
